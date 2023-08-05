from datetime import datetime
from asyncio import ensure_future, get_event_loop
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles.colors import RED, BLUE, DARKGREEN
from openpyxl.styles.fills import PatternFill, FILL_SOLID
from openpyxl.worksheet.dimensions import ColumnDimension
from motor.motor_asyncio import AsyncIOMotorClient

from .database import DbHelper


class OutboundXls:
    def __init__(self):
        self._collection = DbHelper.mongo.contactcenter['outbound-postprocessing']

    async def asxls(self, begin, end):
        """
        :rtype: bytes
        """

        return await get_event_loop().run_in_executor(
            None,
            self._toxls,
            await self._select_data(begin, end)
        )

    async def _select_data(self, begin, end, date_format='%Y-%m-%d'):
        """
        :rtype: list
        :description: just take data from mongo
        """

        return list(dict(
            description=document['description'],
            label=document['label'],
            operator=document['operatorName'],
            fields=document['fields']
        ) for document in await self._collection.find({
            "time": {"$gte": datetime.strptime(begin, date_format)},
            "time": {"$lte": datetime.strptime(end, date_format)}
        }).to_list(100))

    def _toxls(self, documents: list):
        """
        :rtype: bytes
        """

        wb = Workbook()
        sheet = wb.active
        dimensions = sheet.column_dimensions
        row = 1

        for dimension in dimensions['A'], dimensions['B'], dimensions['C']:
            dimension.width = 50

        for document in documents:
            # write document data
            sheet[f"A{row}"] = document['label']
            sheet[f"B{row}"] = document['description']
            sheet[f"C{row}"] = document['operator']

            for cell in sheet[f"A{row}"], sheet[f"B{row}"], sheet[f"C{row}"]:
                cell.fill = PatternFill(start_color="00FF8080", end_color="00FF8080", fill_type=FILL_SOLID)

            row += 1

            # write fields data
            for field in document['fields']:
                sheet[f"A{row}"] = field['label']
                sheet[f"B{row}"] = ", ".join(field['value'])
                sheet[f"C{row}"] = f"[{field['type']}]"

                row += 1

        return save_virtual_workbook(wb)
